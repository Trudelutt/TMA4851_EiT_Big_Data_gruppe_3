import numpy as np
import csv
import pandas as pd
import math
import pycountry
from scipy.stats import norm
import time
from openpyxl import load_workbook
import json

start = time.time()

avs=pd.read_csv('capdist.csv', sep=',',header=None)
avs = avs.values
df=pd.read_csv('TMformat.csv', sep=',',header=None,encoding='latin-1')
df = df.values


def avstand(land1,land2,avs):
    for i in range(len(avs)):
        if(avs[i][0]==land1 and avs[i][1]==land2):
            return avs[i][2]
    return 0
   
    

def alpha3ToCountryName(alpha):
	try:
		country = pycountry.countries.get(alpha3=alpha)
	except KeyError:
		#print("Could not recognise code - default value returned")
		country = pycountry.countries.get(alpha3="ATA")
	countryName = country.name
	return countryName

def countryNameToAlpha3(cName):
	try:
		country = pycountry.countries.get(name=cName)
	except KeyError:
		#print("Could not recognise code - default value returned")
		country = pycountry.countries.get(name="Antarctica")
	alpha = country.alpha3
	return alpha


def TM_vare_import_og_eksport(land1,land2,item,df):
    
    import_Tonn_Array = np.array([])
    eksport_Tonn_Array = np.array([])
    for i in range(len(df)):
        if df[i][1]==land1 and df[i][3]==land2 and df[i][5]==item and df[i][7]=='Import Quantity':
            for k in range(1,56,2):
                import_Tonn_Array = np.append(import_Tonn_Array,df[i][-(k+1)])
        
        if df[i][1]==land1 and df[i][3]==land2 and df[i][5]==item and df[i][7]=='Export Quantity':
            for k in range(1,56,2):
                eksport_Tonn_Array = np.append(eksport_Tonn_Array,df[i][-(k+1)])
            #if df[i+1][1]!=land1:
                #break
    if len(import_Tonn_Array)==0:
       import_Tonn_Array = np.zeros(28)
       
       print('error, making import zero vector')
  
    if len(eksport_Tonn_Array)==0:  
        eksport_Tonn_Array = np.zeros(28)
        print('error, making export zero vector')
        
    maxLen=28
    for i in range(maxLen):
        if math.isnan(import_Tonn_Array[i]):
            import_Tonn_Array[i]=0
        if math.isnan(eksport_Tonn_Array[i]):
            eksport_Tonn_Array[i]=0
    eksport_Tonn_Array = np.flip(eksport_Tonn_Array,0)
    import_Tonn_Array = np.flip(import_Tonn_Array,0)
    return import_Tonn_Array,eksport_Tonn_Array




def indeksOneYear(country,year,df):
    indeksForCountry = 0
    
    #eksport_Tonn_Array=np.array([])
    for i in range(len(df)-1): 
        import_Tonn_Array=np.array([])
        if df[i][1]==country:
          country2 = countryNameToAlpha3(df[i][3])
          if(df[i][7]=='Import Quantity'):
              for k in range(1,29,1):
                import_Tonn_Array = np.append(import_Tonn_Array,df[i][-(k)])  
          if len(import_Tonn_Array)==0:
              import_Tonn_Array = np.zeros(28)
          if df[i+1][1]!=country:
              break
  
          maxLen=28
          for i in range(maxLen):
            if math.isnan(import_Tonn_Array[i]):
                import_Tonn_Array[i]=0
          import_Tonn_Array = np.flip(import_Tonn_Array,0)
          importNumbersYear = import_Tonn_Array[year]
          indeksForCountry += importNumbersYear*punish(country,country2)
                    

    return indeksForCountry


def indeksAllYears(country,df):
    indeksForCountry = np.zeros(28)
    
    #eksport_Tonn_Array=np.array([])
    for i in range(len(df)-1): 
        import_Tonn_Array=np.array([])
        if df[i][1]==country:
          country2 = countryNameToAlpha3(df[i][3])
          if(df[i][7]=='Import Quantity'):
              for k in range(1,29,1):
                import_Tonn_Array = np.append(import_Tonn_Array,df[i][-(k)])
          if len(import_Tonn_Array)==0:
              import_Tonn_Array = np.zeros(28)
          if df[i+1][1]!=country:
              break
  
          maxLen=28
          for i in range(maxLen):
            if math.isnan(import_Tonn_Array[i]):
                import_Tonn_Array[i]=0
          import_Tonn_Array = np.flip(import_Tonn_Array,0)
          
          
          importNumbersYear = import_Tonn_Array
          
          indeksForCountry = [sum(x) for x in zip(indeksForCountry, importNumbersYear*punish(country,country2))] 
                    

    return indeksForCountry


def punish(country1,country2):
    distance = int(avstand(country1,country2,avs))
    punishmentFactor = norm.cdf(distance,7000,5000)
    return punishmentFactor


def importPopulation():
	wb = load_workbook(filename = 'pop_faostat.xlsx', data_only=True)
	sh = wb['population']
	popMatrix = []
	countries = []
	years = 2017-1949

	rows_iter = sh.iter_rows(min_col=1, min_row=2, max_col=4, max_row=sh.max_row)
	vals = [[cell.value for cell in row] for row in rows_iter]
	nrows = len(vals)

	current_id = vals[nrows-1][0] 
	prev_id = -1
	country_pop = []
	for row in range(nrows):
		current_id = vals[row][0] 
		if current_id != prev_id:
			popMatrix.append(country_pop)
			country_name = vals[row][1]
			country_pop = [country_name] + [0 for i in range(years)]
			prev_id = current_id

		year_id = vals[row][2]-1949
		country_pop[year_id] = vals[row][3]

	popMatrix.append(country_pop)
	popMatrix.pop(0)

	for row in range(len(popMatrix)):
		name = popMatrix[row][0]
		alpha3 = countryNameToAlpha3(name)
		if alpha3 == "ATA":
			print("Problem with " + name)
		else:
			popMatrix[row][0] = alpha3
	return popMatrix

def getPopulation(alfa3_country, year, populationMatrix):
	for i in range(len(populationMatrix)):
		if alfa3_country == populationMatrix[i][0]:
			return populationMatrix[i][year-1949]
	return -1


def normIndeks(country,df,pop):
    temp = np.asarray(indeksAllYears(country,df))
    #years = np.array([range(1983,2014,1)])
    pops = []
    alpha3Country = countryNameToAlpha3(country)
    for i in range(1986,2014,1):
        pops.append(getPopulation(alpha3Country,i,pop))
    indeksForCountry = temp/np.asarray(pops)
    
    
    return indeksForCountry

def run(df,pop,avs):
    countryName = [i for i in avs[:,1]]
    #countryName[203:41007]=[]
    countryName[203:41007]=[]
    countryName.pop(0)
    #countryName = ['NOR','DNK']
    #nameList = ['Afghanistan','Austria','Canada','Colombia']
    utMatrix = [[] for i in range((len(countryName)))]
    counter = 0
    for str in countryName:
        longstr = alpha3ToCountryName(str)
        temp = normIndeks(longstr,df,pop)
        utMatrix[counter].append(longstr)
        for i in range(len(temp)):
            utMatrix[counter].append(temp[i])
        counter = counter+1
        print(counter)
    #utMatrix = np.asarray(utMatrix)    
    #np.savetxt('utMatrix.txt', utMatrix, delimiter=',')   # X is an array
    csvfile = 'utMatrix.txt'
    with open(csvfile, "w") as output:
        writer = csv.writer(output, lineterminator='\n')
        writer.writerows(utMatrix)


pop = importPopulation()
start = time.time()
#test = normIndeks('Denmark',df,pop)
#run(df,pop,avs)
end = time.time()
#print(end-start)
